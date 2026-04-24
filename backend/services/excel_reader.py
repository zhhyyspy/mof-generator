"""
V3.4: Excel/CSV raw cell reader for structured table extraction.

Goal: dump the raw cell grid (values + merged cell ranges + basic style info)
so that downstream AI can reason about table structure (where headers live,
where data starts/ends, which rows are summary, which columns are decorative).

Intentionally does NOT try to infer structure — that's Phase 1 (AI).

Output shape:
  {
    "workbook": {"sheets": [sheet_name, ...], "path": "..."},
    "sheets": {
      "Sheet1": {
        "max_row": 125,
        "max_col": 7,
        "merged_ranges": [
          {"min_row": 1, "max_row": 1, "min_col": 1, "max_col": 7, "value": "XX台账"},
          ...
        ],
        "cells": [
          [null, {"v": "序号", "b": true}, ...],  // row 1
          ...
        ]
      }
    }
  }

`cells[r][c]` is None for empty cell, or a dict with:
  - "v": display value (str / number / None)
  - "b": true if bold (heuristic for headers)
  - "m": true if this cell is INSIDE a merged range but not the top-left
"""
from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any, Optional


MAX_PREVIEW_ROWS = 50       # cap for AI analysis
MAX_EXTRACT_ROWS = 5000     # cap for actual data ingestion (M0 sample or full)


def read_xlsx_raw(path: Path, max_rows: int = MAX_PREVIEW_ROWS) -> dict:
    """Read an .xlsx file's raw cell grid without interpretation.

    Returns the structure described at module top. Uses openpyxl in
    read_only mode with data_only=True (so formulas are resolved to cached
    values, not formula strings).
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=False, data_only=True)
    result: dict[str, Any] = {
        "workbook": {"sheets": wb.sheetnames, "path": str(path)},
        "sheets": {},
    }
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # We DO want merged range info, so use regular mode (not read_only)
        max_row = min(ws.max_row or 0, max_rows)
        max_col = ws.max_column or 0

        # Merged ranges (list of openpyxl.worksheet.cell_range.CellRange)
        merged = []
        merged_cells_set: set[tuple[int, int]] = set()   # (row, col) inside a merge
        merged_tops: dict[tuple[int, int], dict] = {}    # top-left → full info
        for mr in ws.merged_cells.ranges:
            if mr.min_row > max_rows:
                continue
            # top-left holds the value
            tl_val = ws.cell(row=mr.min_row, column=mr.min_col).value
            merged.append({
                "min_row": mr.min_row, "max_row": mr.max_row,
                "min_col": mr.min_col, "max_col": mr.max_col,
                "value": _safe_cell_value(tl_val),
            })
            merged_tops[(mr.min_row, mr.min_col)] = merged[-1]
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    if (r, c) != (mr.min_row, mr.min_col):
                        merged_cells_set.add((r, c))

        # Dump cell grid
        cells: list[list[Optional[dict]]] = []
        for r in range(1, max_row + 1):
            row: list[Optional[dict]] = []
            for c in range(1, max_col + 1):
                if (r, c) in merged_cells_set:
                    row.append({"v": None, "m": True})
                    continue
                cell = ws.cell(row=r, column=c)
                v = _safe_cell_value(cell.value)
                if v is None or v == "":
                    row.append(None)
                else:
                    entry = {"v": v}
                    # Bold flag heuristic (headers are often bold)
                    try:
                        if cell.font and cell.font.bold:
                            entry["b"] = True
                    except Exception:
                        pass
                    row.append(entry)
            cells.append(row)

        result["sheets"][sheet_name] = {
            "max_row": max_row,
            "max_col": max_col,
            "merged_ranges": merged,
            "cells": cells,
        }
    wb.close()
    return result


def read_csv_raw(path: Path, max_rows: int = MAX_PREVIEW_ROWS) -> dict:
    """Read a CSV as a single-sheet raw grid. No merged cells possible."""
    rows: list[list[Optional[dict]]] = []
    max_col = 0
    # Detect encoding: try utf-8 first, fall back to GBK (common in CN Excel exports)
    text: Optional[str] = None
    for enc in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            with open(path, "r", encoding=enc) as f:
                text = f.read()
            break
        except (UnicodeDecodeError, Exception):
            continue
    if text is None:
        raise ValueError(f"无法解析 CSV 文件编码: {path}")

    reader = csv.reader(io.StringIO(text))
    for i, row in enumerate(reader):
        if i >= max_rows:
            break
        cleaned: list[Optional[dict]] = []
        for v in row:
            if v is None or v == "":
                cleaned.append(None)
            else:
                cleaned.append({"v": v})
        if len(cleaned) > max_col:
            max_col = len(cleaned)
        rows.append(cleaned)
    # Pad rows to max_col
    for r in rows:
        while len(r) < max_col:
            r.append(None)

    return {
        "workbook": {"sheets": ["Sheet1"], "path": str(path)},
        "sheets": {
            "Sheet1": {
                "max_row": len(rows),
                "max_col": max_col,
                "merged_ranges": [],
                "cells": rows,
            }
        },
    }


def read_structured_file(path: Path, max_rows: int = MAX_PREVIEW_ROWS) -> dict:
    """Dispatch on file extension."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm", ".xls"):
        return read_xlsx_raw(path, max_rows=max_rows)
    if suffix == ".csv":
        return read_csv_raw(path, max_rows=max_rows)
    raise ValueError(f"不支持的表格文件类型: {suffix}")


def is_structured_file(filename: str) -> bool:
    """Check if a filename looks like a structured data file (not doc)."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return ext in ("xlsx", "xlsm", "xls", "csv")


def _safe_cell_value(v: Any) -> Any:
    """Coerce openpyxl cell values into JSON-safe primitives.
    Datetime → ISO string, etc."""
    from datetime import datetime, date, time
    if v is None:
        return None
    if isinstance(v, (datetime, date, time)):
        return v.isoformat()
    if isinstance(v, (int, float, str, bool)):
        return v
    return str(v)


def extract_table_data(
    path: Path,
    sheet_name: str,
    header_rows: list[int],
    data_start: int,
    data_end: int,
    summary_rows: Optional[list[int]] = None,
    ignored_cols: Optional[list[int]] = None,
    selected_cols: Optional[list[int]] = None,
    max_sample: int = 20,
) -> list[dict[str, Any]]:
    """Given user-confirmed table boundaries, extract data rows as dicts.

    Used for:
     - M0 sample storage (first 5-20 rows)
     - (future) full M0 ingestion if we ever support it

    Each row returned keyed by the DERIVED COLUMN NAME (from header row / joined
    multi-level header if applicable).

    Column naming: if multiple header_rows given, concatenate them with " · ".
    E.g. "设备基本信息 · 型号"  — business-friendly key.
    """
    raw = read_structured_file(path, max_rows=max(data_end, max_sample + 100))
    sheet_data = raw["sheets"].get(sheet_name)
    if sheet_data is None:
        raise ValueError(f"Sheet '{sheet_name}' not found")
    cells = sheet_data["cells"]
    max_col = sheet_data["max_col"]
    ignored = set(ignored_cols or [])
    summary = set(summary_rows or [])

    # Build column name per column index (1-based)
    col_names: dict[int, str] = {}
    for c in range(1, max_col + 1):
        if c in ignored:
            continue
        parts: list[str] = []
        for hr in header_rows:
            if 0 < hr <= len(cells):
                cell = cells[hr - 1][c - 1] if c <= len(cells[hr - 1]) else None
                if cell and cell.get("v"):
                    s = str(cell["v"]).strip()
                    if s and (not parts or parts[-1] != s):
                        parts.append(s)
        if parts:
            col_names[c] = " · ".join(parts)

    if selected_cols:
        col_names = {c: n for c, n in col_names.items() if c in set(selected_cols)}

    # Extract data rows
    rows_out: list[dict[str, Any]] = []
    for r in range(data_start, min(data_end, len(cells)) + 1):
        if r in summary:
            continue
        row_cells = cells[r - 1] if r - 1 < len(cells) else []
        row_dict: dict[str, Any] = {}
        for c, col_name in col_names.items():
            cell = row_cells[c - 1] if c - 1 < len(row_cells) else None
            row_dict[col_name] = cell.get("v") if cell else None
        # Skip entirely empty rows
        if any(v is not None and v != "" for v in row_dict.values()):
            rows_out.append(row_dict)
        if len(rows_out) >= max_sample:
            break
    return rows_out
